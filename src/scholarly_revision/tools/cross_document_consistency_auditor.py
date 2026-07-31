'''Audit Phase 7 consistency across JSON, DOCX, XLSX, and registries.'''

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scholarly_revision.models.release import (
    ConsistencyCategory, ConsistencyFinding,
)
from scholarly_revision.models.response_package import (
    CommentResolution, ResponsePackage, ResponseStatus,
)
from scholarly_revision.models.scientific_audit import AuditSeverity
from scholarly_revision.services.gap_analysis_service import read_json
from scholarly_revision.services.response_letter_service import load_response_sources
from scholarly_revision.tools.docx_clean_copy import validate_text_equivalence
from scholarly_revision.tools.docx_reader import read_docx
from scholarly_revision.tools.location_verifier import verify_locations
from scholarly_revision.tools.response_docx_builder import response_docx_entry_records


@dataclass(frozen=True, slots=True)
class ConsistencyAuditReport:
    generated_at: datetime
    findings: tuple[ConsistencyFinding, ...]
    passed: bool
    count_by_category: dict[str, int]
    count_by_severity: dict[str, int]

    def to_dict(self) -> dict[str, Any]:
        return {
            'schema_version': 1,
            'generated_at': self.generated_at.isoformat(),
            'passed': self.passed,
            'finding_count': len(self.findings),
            'count_by_category': self.count_by_category,
            'count_by_severity': self.count_by_severity,
            'findings': [item.model_dump(mode='json') for item in self.findings],
        }


class _Findings:
    def __init__(self) -> None:
        self.items: list[ConsistencyFinding] = []

    def add(
        self, category: ConsistencyCategory, severity: AuditSeverity,
        description: str, *, comments: list[str] | None = None,
        actions: list[str] | None = None, changes: list[str] | None = None,
        responses: list[str] | None = None, documents: list[str] | None = None,
        details: dict[str, object] | None = None,
    ) -> None:
        self.items.append(ConsistencyFinding(
            finding_id=f'CONS-{len(self.items) + 1:04d}',
            category=category, severity=severity, description=description,
            related_comment_ids=comments or [], related_action_ids=actions or [],
            related_change_ids=changes or [], related_response_entry_ids=responses or [],
            documents=documents or [], details=details or {},
        ))


def _package(root: Path, package: ResponsePackage | dict | str | Path | None) -> ResponsePackage:
    if package is None:
        package = read_json(root / 'working' / 'response_package.json')
    elif isinstance(package, (str, Path)):
        package = read_json(package)
    if isinstance(package, ResponsePackage):
        package = package.model_dump(mode='python')
    return ResponsePackage.model_validate(package)


def _text(path: Path) -> str:
    return '\n'.join(item.text for item in read_docx(path))


def audit_cross_document_consistency(
    project_root: str | Path,
    package: ResponsePackage | dict | str | Path | None = None,
    *,
    response_letter: str | Path | None = None,
) -> ConsistencyAuditReport:
    root = Path(project_root).expanduser().resolve()
    response = _package(root, package)
    sources = load_response_sources(root)
    comments = {item.comment_id: item for item in sources['comments']}
    actions = {item.action_id: item for item in sources['actions']}
    changes = {item.change_id: item for item in sources['changes']}
    entries = response.entries
    by_comment = {item.comment_id: item for item in entries}
    evidence = {str(item.get('evidence_id')): item for item in sources['evidence']}
    references = {str(item.get('reference_id')): item for item in sources['references']}
    found = _Findings()

    response_counts = Counter(item.comment_id for item in entries)
    for comment_id in sorted(set(comments) - set(by_comment)):
        found.add(
            ConsistencyCategory.COMMENT_COVERAGE, AuditSeverity.BLOCKER,
            'Reviewer comment is missing from the response letter.',
            comments=[comment_id], documents=['reviewer_comments.json', 'response_package.json'],
        )
    for comment_id, count in sorted(response_counts.items()):
        if count > 1:
            found.add(
                ConsistencyCategory.COMMENT_COVERAGE, AuditSeverity.BLOCKER,
                'Reviewer comment has duplicated response entries.',
                comments=[comment_id], details={'count': count},
            )
    for entry in entries:
        comment = comments.get(entry.comment_id)
        if comment is None:
            found.add(
                ConsistencyCategory.COMMENT_COVERAGE, AuditSeverity.BLOCKER,
                'Response entry has no reviewer comment record.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )
        elif entry.exact_comment != comment.original_comment:
            found.add(
                ConsistencyCategory.TRACEABILITY, AuditSeverity.BLOCKER,
                'Response entry does not preserve the exact reviewer comment.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )
        if comment is not None and entry.highlight is not comment.highlight:
            found.add(
                ConsistencyCategory.HIGHLIGHT, AuditSeverity.CRITICAL,
                'Response uses the wrong reviewer highlight.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )
        if entry.resolution is None:
            found.add(
                ConsistencyCategory.STATUS, AuditSeverity.BLOCKER,
                'Reviewer comment has no explicit response state.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )

    logged_actions = {item.action_id for item in changes.values()}
    for action in actions.values():
        if action.approval_state.value == 'APPROVED' and action.action_id not in logged_actions:
            found.add(
                ConsistencyCategory.TRACEABILITY, AuditSeverity.CRITICAL,
                'Approved action is missing from Change Log.',
                comments=action.comment_ids, actions=[action.action_id],
                documents=['revision_plan.json', 'change_log.json'],
            )
    response_changes = {
        change_id for entry in entries for change_id in entry.related_change_ids
    }
    for change in changes.values():
        if change.change_id not in response_changes:
            found.add(
                ConsistencyCategory.CHANGE_CLAIM, AuditSeverity.CRITICAL,
                'Applied change is absent from the response.',
                comments=change.comment_ids, actions=[change.action_id],
                changes=[change.change_id],
            )
    highlighted = root / 'outputs' / 'Revised_Manuscript_Highlighted.docx'
    clean = root / 'outputs' / 'Revised_Manuscript_Clean.docx'
    manuscript_text = _text(highlighted) if highlighted.is_file() else ''
    page_path = root / 'audit' / 'rendered_location_metadata.json'
    page_metadata = read_json(page_path) if page_path.is_file() else None
    for entry in entries:
        for change_id in entry.related_change_ids:
            change = changes.get(change_id)
            if change is None:
                found.add(
                    ConsistencyCategory.CHANGE_CLAIM, AuditSeverity.BLOCKER,
                    'Response claims an unapplied revision.',
                    comments=[entry.comment_id], changes=[change_id],
                    responses=[entry.response_entry_id],
                )
                continue
            action = actions.get(change.action_id)
            if action is None or action.approval_state.value != 'APPROVED':
                found.add(
                    ConsistencyCategory.APPROVAL, AuditSeverity.BLOCKER,
                    'An unapproved revision was applied or reported.',
                    comments=[entry.comment_id], actions=[change.action_id],
                    changes=[change_id],
                )
            if action is not None and change.highlight is not action.highlight:
                found.add(
                    ConsistencyCategory.HIGHLIGHT, AuditSeverity.CRITICAL,
                    'Change Log highlight conflicts with the approved action policy.',
                    comments=[entry.comment_id], changes=[change_id],
                    responses=[entry.response_entry_id],
                )
        for evidence_id in entry.related_evidence_ids:
            item = evidence.get(evidence_id)
            if item is None or str(item.get('status')) != 'VERIFIED':
                found.add(
                    ConsistencyCategory.EVIDENCE, AuditSeverity.BLOCKER,
                    'Response references missing or unverified evidence.',
                    comments=[entry.comment_id], responses=[entry.response_entry_id],
                    details={'evidence_id': evidence_id},
                )
        for reference_id in entry.related_reference_ids:
            if reference_id not in references:
                found.add(
                    ConsistencyCategory.REFERENCE, AuditSeverity.BLOCKER,
                    'Response references a missing bibliography record.',
                    comments=[entry.comment_id], responses=[entry.response_entry_id],
                    details={'reference_id': reference_id},
                )
        if entry.verified_locations and highlighted.is_file():
            results = verify_locations(
                highlighted, entry.verified_locations, page_metadata=page_metadata
            )
            if any(not item.verified for item in results):
                found.add(
                    ConsistencyCategory.LOCATION, AuditSeverity.CRITICAL,
                    'Response contains an inconsistent or unverified location.',
                    comments=[entry.comment_id], responses=[entry.response_entry_id],
                )
        elif entry.related_change_ids:
            found.add(
                ConsistencyCategory.LOCATION, AuditSeverity.CRITICAL,
                'Response for an applied change has no verified location.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )
        numeric = set(re.findall(r'(?<![A-Za-z-])\d+\.\d+%?', entry.changes_made))
        missing_numbers = sorted(value for value in numeric if value not in manuscript_text)
        if missing_numbers:
            found.add(
                ConsistencyCategory.NUMERICAL, AuditSeverity.BLOCKER,
                'Numerical value in the response conflicts with or is absent from the manuscript.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
                details={'values': missing_numbers},
            )
        linked = [
            action for action in actions.values() if entry.comment_id in action.comment_ids
        ]
        deferred = any(
            action.status.value == 'DEFERRED'
            or action.approval_decision in {'DEFER', 'NEED_MORE_EVIDENCE'}
            for action in linked
        )
        if deferred and entry.resolution in {
            CommentResolution.FULLY_ADDRESSED,
            CommentResolution.PARTIALLY_ADDRESSED,
        }:
            found.add(
                ConsistencyCategory.STATUS, AuditSeverity.BLOCKER,
                'Comment is marked complete despite a deferred or evidence-blocked action.',
                comments=[entry.comment_id], responses=[entry.response_entry_id],
            )
    if not highlighted.is_file() or not clean.is_file():
        found.add(
            ConsistencyCategory.ARTIFACT, AuditSeverity.BLOCKER,
            'Clean or highlighted revised manuscript is missing.',
            documents=['Revised_Manuscript_Highlighted.docx', 'Revised_Manuscript_Clean.docx'],
        )
    else:
        try:
            equivalent = validate_text_equivalence(highlighted, clean)
        except Exception as exc:
            equivalent = False
            found.add(
                ConsistencyCategory.ARTIFACT, AuditSeverity.BLOCKER,
                'Clean or highlighted manuscript cannot be opened as a valid DOCX.',
                documents=[highlighted.name, clean.name],
                details={'error_type': type(exc).__name__},
            )
        if not equivalent:
            found.add(
                ConsistencyCategory.MANUSCRIPT_EQUIVALENCE, AuditSeverity.BLOCKER,
                'Clean and highlighted manuscript text does not match.',
                documents=[highlighted.name, clean.name],
            )
    qa = sources['qa']
    for issue in qa.get('issues', []) if isinstance(qa, dict) else []:
        if (
            str(issue.get('severity')) not in {'BLOCKER', 'CRITICAL'}
            or str(issue.get('status', 'OPEN')) not in {'OPEN', 'ACKNOWLEDGED'}
        ):
            continue
        for comment_id in issue.get('related_comment_ids', []):
            entry = by_comment.get(comment_id)
            if entry is not None and entry.resolution not in {
                CommentResolution.DEFERRED,
                CommentResolution.BLOCKED_BY_MISSING_EVIDENCE,
            }:
                found.add(
                    ConsistencyCategory.QA, AuditSeverity.BLOCKER,
                    'Unresolved QA blocker is omitted from the response.',
                    comments=[comment_id], responses=[entry.response_entry_id],
                    details={'qa_issue_id': issue.get('issue_id')},
                )
    letter = Path(response_letter) if response_letter else root / 'outputs' / 'Response_to_Reviewers.docx'
    if letter.is_file():
        try:
            records = response_docx_entry_records(letter)
        except Exception as exc:
            records = ()
            found.add(
                ConsistencyCategory.ARTIFACT, AuditSeverity.BLOCKER,
                'Response DOCX structure cannot be validated.',
                documents=[letter.name],
                details={'error_type': type(exc).__name__},
            )
        if len(records) != len(entries):
            found.add(
                ConsistencyCategory.COMMENT_COVERAGE, AuditSeverity.BLOCKER,
                'Response DOCX entry count differs from response_package.json.',
                documents=[letter.name, 'response_package.json'],
                details={
                    'docx_entry_count': len(records),
                    'package_entry_count': len(entries),
                },
            )
        for entry, record in zip(entries, records):
            expected_heading = (
                f'Reviewer {entry.reviewer_number}, Comment {entry.sequence_number}'
                if entry.reviewer_source.value == 'REVIEWER' else
                f'Editor, Comment {entry.sequence_number}'
                if entry.reviewer_source.value == 'EDITOR' else
                f'General Comment {entry.sequence_number}'
            )
            comparisons = (
                (
                    'heading', record.heading, expected_heading,
                    ConsistencyCategory.TRACEABILITY,
                ),
                (
                    'exact comment', record.comment, entry.exact_comment,
                    ConsistencyCategory.COMMENT_COVERAGE,
                ),
                (
                    'author response', record.author_response,
                    entry.author_response or 'Response pending.',
                    ConsistencyCategory.TRACEABILITY,
                ),
                (
                    'changes made', record.changes_made,
                    entry.changes_made or 'No manuscript change reported.',
                    ConsistencyCategory.CHANGE_CLAIM,
                ),
                (
                    'location', record.location,
                    '; '.join(entry.verified_locations) or 'Not required.',
                    ConsistencyCategory.LOCATION,
                ),
                (
                    'highlight label', record.highlight,
                    entry.highlight.value.replace('_', ' ').title(),
                    ConsistencyCategory.HIGHLIGHT,
                ),
            )
            for field, actual, expected, category in comparisons:
                if actual == expected:
                    continue
                found.add(
                    category, AuditSeverity.BLOCKER,
                    f'Response DOCX {field} differs from response_package.json.',
                    comments=[entry.comment_id], responses=[entry.response_entry_id],
                    documents=[letter.name],
                )
            if record.heading_highlight is not entry.highlight:
                found.add(
                    ConsistencyCategory.HIGHLIGHT, AuditSeverity.CRITICAL,
                    'Response DOCX visible heading highlight conflicts with policy.',
                    comments=[entry.comment_id], responses=[entry.response_entry_id],
                    documents=[letter.name],
                )
    else:
        found.add(
            ConsistencyCategory.ARTIFACT, AuditSeverity.BLOCKER,
            'Response-to-reviewers DOCX is missing.',
            documents=['Response_to_Reviewers.docx'],
        )
    workbook_path = root / 'outputs' / 'Revision_Master.xlsx'
    if workbook_path.is_file():
        workbook = load_workbook(workbook_path, data_only=False)
        sheet = workbook['Response_Map']
        headers = {
            str(cell.value): cell.column for cell in sheet[1] if cell.value is not None
        }
        comment_column = headers.get('Comment ID')
        status_column = headers.get('Response Status') or headers.get('Status')
        verification_column = headers.get('Verification Status')
        if comment_column and status_column:
            rows = {
                str(sheet.cell(row, comment_column).value): row
                for row in range(2, sheet.max_row + 1)
                if sheet.cell(row, comment_column).value
            }
            for entry in entries:
                row = rows.get(entry.comment_id)
                if row is None:
                    found.add(
                        ConsistencyCategory.STATUS, AuditSeverity.CRITICAL,
                        'Workbook Response_Map is missing a response entry.',
                        comments=[entry.comment_id], documents=[workbook_path.name],
                    )
                    continue
                workbook_status = str(sheet.cell(row, status_column).value or '')
                if workbook_status and workbook_status != entry.response_status.value:
                    found.add(
                        ConsistencyCategory.STATUS, AuditSeverity.CRITICAL,
                        'Response status differs between workbook and JSON.',
                        comments=[entry.comment_id], responses=[entry.response_entry_id],
                        documents=[workbook_path.name, 'response_package.json'],
                        details={
                            'workbook_status': workbook_status,
                            'json_status': entry.response_status.value,
                        },
                    )
                if verification_column:
                    verified = str(sheet.cell(row, verification_column).value or '')
                    if entry.response_status is ResponseStatus.VERIFIED and verified not in {
                        'VERIFIED', 'True', 'TRUE',
                    }:
                        found.add(
                            ConsistencyCategory.STATUS, AuditSeverity.CRITICAL,
                            'Workbook verification status conflicts with verified response JSON.',
                            comments=[entry.comment_id],
                        )
                expected_cells = {
                    'Response Entry ID': entry.response_entry_id,
                    'Change IDs': '; '.join(entry.related_change_ids),
                    'Evidence IDs': '; '.join(entry.related_evidence_ids),
                    'Reference IDs': '; '.join(entry.related_reference_ids),
                    'Verified Location': '; '.join(entry.verified_locations),
                    'Highlight': entry.highlight.value,
                    'Author Approval': entry.author_approved,
                    'Response Draft': entry.author_response,
                    'Changes Made': entry.changes_made,
                    'Applied Location': '; '.join(entry.verified_locations),
                    'Exact Comment Verified': True,
                }
                for header, expected in expected_cells.items():
                    column = headers.get(header)
                    if column is None:
                        continue
                    actual = sheet.cell(row, column).value
                    if isinstance(expected, bool):
                        matches = actual is expected or str(actual).upper() == str(expected).upper()
                    else:
                        matches = str(actual or '') == expected
                    if matches:
                        continue
                    category = (
                        ConsistencyCategory.HIGHLIGHT if header == 'Highlight'
                        else ConsistencyCategory.LOCATION
                        if header in {'Verified Location', 'Applied Location'}
                        else ConsistencyCategory.CHANGE_CLAIM
                        if header in {'Change IDs', 'Changes Made'}
                        else ConsistencyCategory.EVIDENCE
                        if header == 'Evidence IDs'
                        else ConsistencyCategory.REFERENCE
                        if header == 'Reference IDs'
                        else ConsistencyCategory.APPROVAL
                        if header == 'Author Approval'
                        else ConsistencyCategory.TRACEABILITY
                    )
                    found.add(
                        category, AuditSeverity.CRITICAL,
                        f'Workbook {header} differs from response_package.json.',
                        comments=[entry.comment_id],
                        responses=[entry.response_entry_id],
                        documents=[workbook_path.name, 'response_package.json'],
                    )
        reviewer = workbook['Reviewer_Comments']
        reviewer_headers = {
            str(cell.value): cell.column
            for cell in reviewer[1] if cell.value is not None
        }
        reviewer_rows = {
            str(reviewer.cell(row, reviewer_headers['Comment ID']).value): row
            for row in range(2, reviewer.max_row + 1)
            if reviewer.cell(row, reviewer_headers['Comment ID']).value
        }
        for comment_id, comment in comments.items():
            row = reviewer_rows.get(comment_id)
            if row is None or reviewer.cell(
                row, reviewer_headers['Original Comment']
            ).value != comment.original_comment:
                found.add(
                    ConsistencyCategory.TRACEABILITY, AuditSeverity.BLOCKER,
                    'Workbook does not preserve the exact reviewer comment.',
                    comments=[comment_id],
                    documents=[workbook_path.name, 'reviewer_comments.json'],
                )
        workbook.close()
    else:
        found.add(
            ConsistencyCategory.ARTIFACT, AuditSeverity.BLOCKER,
            'Revision_Master.xlsx is missing.',
            documents=['Revision_Master.xlsx'],
        )
    categories = Counter(item.category.value for item in found.items)
    severities = Counter(item.severity.value for item in found.items)
    passed = not any(item.blocking for item in found.items)
    return ConsistencyAuditReport(
        generated_at=datetime.now(UTC),
        findings=tuple(found.items),
        passed=passed,
        count_by_category=dict(sorted(categories.items())),
        count_by_severity=dict(sorted(severities.items())),
    )
