'''Build the macro-free revision source-of-truth workbook with openpyxl.'''

from __future__ import annotations

from collections.abc import Iterable, Mapping
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

from scholarly_revision.models.enums import (
    ApprovalDecision,
    ApprovalGateStatus,
    CoverageStatus,
    ApprovalState,
    ChangeType,
    CommentCategory,
    CommentPriority,
    EvidenceStatus,
    HighlightColor,
    RevisionOperation,
    RevisionStatus,
)
from scholarly_revision.models.reviewer import ReviewerComment
from scholarly_revision.models.gap_analysis import GapAnalysisAssessment
from scholarly_revision.models.reviewer import RevisionAction
from scholarly_revision.models.revision_draft import ChangeRecord, RevisionDraft
from scholarly_revision.models.release import ConsistencyFinding
from scholarly_revision.models.response_package import ResponsePackage


REVISION_WORKBOOK_SHEETS = (
    'Dashboard',
    'Project_Info',
    'Reviewer_Comments',
    'Revision_Plan',
    'Change_Log',
    'Reference_Audit',
    'Figures_Tables',
    'Notation_Equations',
    'Results_Integrity',
    'Response_Map',
    'QA_Findings',
)

REVIEWER_COMMENT_HEADERS = (
    'Comment ID',
    'Source',
    'Reviewer Number',
    'Sequence',
    'Original Comment',
    'Normalized Comment',
    'Category',
    'Priority',
    'Interpretation',
    'Required Actions',
    'Target Sections',
    'Shared With',
    'Status',
    'Highlight',
    'Evidence Status',
    'Author Decision',
    'Notes',
    'Manual Review Required',
    'Coverage Status',
    'Manuscript Evidence',
    'Missing Elements',
    'Required References',
    'Required Experiments',
    'Required Statistics',
    'Author Decision Required',
    'Gap Analysis Confidence',
    'Approved Draft Count',
    'Applied Change Count',
    'Remaining Unresolved Actions',
)

REVISION_PLAN_HEADERS = (
    'Action ID',
    'Comment IDs',
    'Change Type',
    'Target Section',
    'Target Object',
    'Old Text Summary',
    'Proposed Revision',
    'Rationale',
    'Evidence IDs',
    'Status',
    'Approval State',
    'Highlight',
    'Applied Location',
    'Verified By',
    'Verified At',

    'Evidence Requirements',
    'Reference Requirements',
    'Experiment Requirements',
    'Statistical Requirements',
    'Unresolved Questions',
    'Approval Decision',
    'Author Note',
    'Modified Action Text',
    'Evidence Request',
    'Decision Timestamp',
    'Decision Maker',
    'Draft Status',
    'Exact Text Approval State',
    'Application Status',
    'Output Version',
    'Verified Location',
)
_OTHER_HEADERS = {
    'Project_Info': ('Field', 'Value'),
    'Change_Log': (
        'Change ID', 'Draft ID', 'Action ID', 'Comment IDs', 'Operation',
        'Target Section', 'Target Element ID', 'Old Text Hash',
        'New Text Hash', 'Old Text Summary', 'New Text Summary',
        'Highlight', 'Applied At', 'Output Version', 'Verification Status',
        'Warnings',
    ),
    'Reference_Audit': (
        'Reference ID', 'Comment IDs', 'Temporary Number', 'Final Number',
        'Title', 'Source', 'DOI', 'Bibliographic Verified',
        'Claim Support Verified', 'First Citation Location', 'Highlight', 'Notes',
    ),
    'Figures_Tables': (
        'Object ID', 'Object Type', 'Comment IDs', 'Caption', 'Source File',
        'Numbering Verified', 'Cross References Verified', 'Visual QA Status',
        'Highlight', 'Notes',
    ),
    'Notation_Equations': (
        'Object ID', 'Object Type', 'Comment IDs', 'Notation or Equation',
        'Definition Location', 'Numbering Verified', 'Consistency Verified',
        'Highlight', 'Notes',
    ),
    'Results_Integrity': (
        'Result ID', 'Comment IDs', 'Metric', 'Value', 'Unit', 'Source File',
        'Source Location', 'Result Status', 'Evidence Status', 'Approval State',
        'Verified By', 'Verified At', 'Notes',
    ),
    'Response_Map': (
        'Response ID', 'Comment ID', 'Exact Comment Verified',
        'Response Draft', 'Changes Made', 'Applied Location', 'Status',
        'Highlight', 'Consistency Verified', 'Notes',
    ),
    'QA_Findings': (
        'Finding ID', 'Category', 'Severity', 'Description', 'File', 'Page',
        'Section', 'Object ID', 'Status', 'Resolution', 'Verified',
    ),
}

FINAL_RESPONSE_MAP_HEADERS = (
    'Response Entry ID', 'Comment ID', 'Response Status', 'Change IDs',
    'Evidence IDs', 'Reference IDs', 'Verified Location', 'Highlight',
    'Author Approval', 'Verification Status', 'Unresolved Issues',
)

_HIGHLIGHT_HEX = {
    HighlightColor.YELLOW.value: 'FFFFFF00',
    HighlightColor.BRIGHT_GREEN.value: 'FF00FF00',
    HighlightColor.VIOLET.value: 'FFEE82EE',
}
_HEADER_FILL = PatternFill('solid', fgColor='D9E1F2')
_ENUM_VALUES = {
    'PriorityValues': [item.value for item in CommentPriority],
    'CategoryValues': [item.value for item in CommentCategory],
    'StatusValues': [item.value for item in RevisionStatus],
    'EvidenceStatusValues': [item.value for item in EvidenceStatus],
    'ApprovalStateValues': [item.value for item in ApprovalState],
    'HighlightValues': [item.value for item in HighlightColor],
    'ChangeTypeValues': [item.value for item in ChangeType],
    'OperationValues': [item.value for item in RevisionOperation],


    'CoverageStatusValues': [item.value for item in CoverageStatus],
    'ApprovalDecisionValues': [item.value for item in ApprovalDecision],
}
def _join(values: Iterable[Any]) -> str:
    return '; '.join(str(value.value if hasattr(value, 'value') else value) for value in values)


def _format_header(worksheet: Any, headers: tuple[str, ...]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = worksheet.cell(1, column, header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    worksheet.freeze_panes = 'A2'
    worksheet.auto_filter.ref = f'A1:{get_column_letter(len(headers))}{max(2, worksheet.max_row)}'
    worksheet.row_dimensions[1].height = 30


def _format_body(worksheet: Any, headers: tuple[str, ...]) -> None:
    long_columns = {
        'Original Comment', 'Normalized Comment', 'Interpretation',
        'Required Actions', 'Target Sections', 'Notes', 'Proposed Revision',
        'Rationale', 'Old Text Summary', 'Description', 'Response Draft',
        'Changes Made', 'Resolution', 'Caption',
    }
    for column, header in enumerate(headers, start=1):
        width = 42 if header in long_columns else max(13, min(26, len(header) + 3))
        worksheet.column_dimensions[get_column_letter(column)].width = width
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)


def _write_enum_lists(workbook: Workbook) -> None:
    sheet = workbook['Project_Info']
    start_column = 24
    for offset, (name, values) in enumerate(_ENUM_VALUES.items()):
        column = start_column + offset
        letter = get_column_letter(column)
        sheet.cell(1, column, name)
        for row, value in enumerate(values, start=2):
            sheet.cell(row, column, value)
        sheet.column_dimensions[letter].hidden = True
        reference = f'{quote_sheetname(sheet.title)}!${letter}$2:${letter}${len(values) + 1}'
        workbook.defined_names.add(DefinedName(name, attr_text=reference))


def _add_validation(worksheet: Any, column: int, enum_name: str) -> None:
    validation = DataValidation(type='list', formula1=f'={enum_name}', allow_blank=True)
    validation.error = f'Choose a value from {enum_name}.'
    validation.errorTitle = 'Invalid controlled value'
    validation.promptTitle = 'Controlled value'
    validation.prompt = 'Select a repository-approved value.'
    validation.showErrorMessage = True
    validation.showInputMessage = True
    worksheet.add_data_validation(validation)
    letter = get_column_letter(column)
    validation.add(f'{letter}2:{letter}1000')


def _add_validations(workbook: Workbook) -> None:
    reviewer = workbook['Reviewer_Comments']
    for column, name in {
        7: 'CategoryValues', 8: 'PriorityValues', 13: 'StatusValues',
        14: 'HighlightValues', 15: 'EvidenceStatusValues',
        19: 'CoverageStatusValues',
    }.items():
        _add_validation(reviewer, column, name)
    plan = workbook['Revision_Plan']
    for column, name in {
        3: 'ChangeTypeValues', 10: 'StatusValues',
        11: 'ApprovalStateValues', 12: 'HighlightValues',
        21: 'ApprovalDecisionValues',
    }.items():
        _add_validation(plan, column, name)
    for sheet_name, column, name in (
        ('Change_Log', 5, 'OperationValues'),
        ('Change_Log', 12, 'HighlightValues'),
        ('Change_Log', 15, 'StatusValues'),
        ('Reference_Audit', 11, 'HighlightValues'),
        ('Figures_Tables', 9, 'HighlightValues'),
        ('Notation_Equations', 8, 'HighlightValues'),
        ('Results_Integrity', 9, 'EvidenceStatusValues'),
        ('Results_Integrity', 10, 'ApprovalStateValues'),
        ('Response_Map', 7, 'StatusValues'),
        ('Response_Map', 8, 'HighlightValues'),
    ):
        _add_validation(workbook[sheet_name], column, name)


def _reviewer_row(comment: ReviewerComment) -> list[Any]:
    return [
        comment.comment_id,
        comment.reviewer_source.value,
        comment.reviewer_number,
        comment.sequence_number,
        comment.original_comment,
        comment.normalized_comment,
        _join(comment.categories),
        comment.priority.value,
        comment.interpretation,
        _join(comment.required_actions),
        _join(comment.target_sections),
        _join(comment.shared_with),
        comment.status.value,
        comment.highlight.value if comment.highlight else None,
        comment.evidence_status.value,
        comment.author_decision,
        comment.notes,
        comment.manual_review_required,
    ]


def _build_dashboard(workbook: Workbook) -> None:
    sheet = workbook['Dashboard']
    sheet.append(['Metric', 'Value'])
    metrics = [
        ('Total comment count', '=COUNTA(Reviewer_Comments!A2:A1000)'),
        ('Count by Reviewer 1', '=COUNTIFS(Reviewer_Comments!B2:B1000,"REVIEWER",Reviewer_Comments!C2:C1000,1)'),
        ('Count by Reviewer 2', '=COUNTIFS(Reviewer_Comments!B2:B1000,"REVIEWER",Reviewer_Comments!C2:C1000,2)'),
        ('Count by Editor/General', '=COUNTIF(Reviewer_Comments!B2:B1000,"EDITOR")+COUNTIF(Reviewer_Comments!B2:B1000,"GENERAL")'),
        ('Count requiring manual review', '=COUNTIF(Reviewer_Comments!R2:R1000,TRUE)'),
        ('Completion percentage', '=IF(B2=0,0,COUNTIF(Reviewer_Comments!M2:M1000,"VERIFIED")/B2)'),
        ('Fully addressed', '=COUNTIF(Reviewer_Comments!S2:S1000,"FULLY_ADDRESSED")'),
        ('Partially addressed', '=COUNTIF(Reviewer_Comments!S2:S1000,"PARTIALLY_ADDRESSED")'),
        ('Not addressed', '=COUNTIF(Reviewer_Comments!S2:S1000,"NOT_ADDRESSED")'),
        ('Cannot determine', '=COUNTIF(Reviewer_Comments!S2:S1000,"CANNOT_DETERMINE")'),
        ('Pending author approval', '=COUNTIF(Revision_Plan!K2:K1000,"PENDING")'),
        ('Approved actions', '=COUNTIF(Revision_Plan!K2:K1000,"APPROVED")'),
        ('Evidence-dependent actions', '=COUNTIF(Revision_Plan!P2:P1000,"?*")'),
        ('Experiment-dependent actions', '=COUNTIF(Revision_Plan!R2:R1000,"?*")'),
        ('Approval Gate status', 'NOT_READY'),
        ('Drafts prepared', 0),
        ('Drafts awaiting text approval', 0),
        ('Text approvals completed', 0),
        ('Changes applied', 0),
        ('Changes blocked', 0),
        ('Document verification status', 'NOT_RUN'),
    ]
    for metric in metrics:
        sheet.append(metric)
    sheet['B7'].number_format = '0.0%'

    sheet.append([])
    sheet.append(['Count by status', 'Count'])
    for status in RevisionStatus:
        row = sheet.max_row + 1
        sheet.append([status.value, f'=COUNTIF(Reviewer_Comments!M2:M1000,A{row})'])
    sheet.append([])
    sheet.append(['Count by priority', 'Count'])
    for priority in CommentPriority:
        row = sheet.max_row + 1
        sheet.append([priority.value, f'=COUNTIF(Reviewer_Comments!H2:H1000,A{row})'])
    sheet.append([])
    sheet.append(['Count by category', 'Count'])
    for category in CommentCategory:
        row = sheet.max_row + 1
        sheet.append([category.value, f'=COUNTIF(Reviewer_Comments!G2:G1000,"*"&A{row}&"*")'])

    sheet['D1'] = 'Highlight Legend'
    sheet['E1'] = 'Assignment'
    for cell in (sheet['D1'], sheet['E1']):
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
    legend = (
        ('YELLOW', 'Reviewer 1'),
        ('BRIGHT_GREEN', 'Reviewer 2'),
        ('VIOLET', 'Editor, general, and shared'),
    )
    for row, (highlight, assignment) in enumerate(legend, start=2):
        sheet.cell(row, 4, highlight).fill = PatternFill('solid', fgColor=_HIGHLIGHT_HEX[highlight])
        sheet.cell(row, 5, assignment)
    sheet.column_dimensions['A'].width = 34
    sheet.column_dimensions['B'].width = 18
    sheet.column_dimensions['D'].width = 22
    sheet.column_dimensions['E'].width = 32
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                cell.protection = Protection(locked=True)
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = f'A1:B{sheet.max_row}'
    sheet.protection.sheet = True


def build_revision_workbook(
    path: str | Path,
    comments: Iterable[ReviewerComment],
    project_info: Mapping[str, Any] | None = None,
) -> Path:
    '''Create ``Revision_Master.xlsx`` as the draft source of truth.'''

    destination = Path(path)
    if destination.suffix.lower() != '.xlsx':
        raise ValueError('revision workbook path must end with .xlsx')
    validated_comments = [ReviewerComment.model_validate(item) for item in comments]

    workbook = Workbook()
    workbook.remove(workbook.active)
    for name in REVISION_WORKBOOK_SHEETS:
        workbook.create_sheet(name)

    project_sheet = workbook['Project_Info']
    project_sheet.append(_OTHER_HEADERS['Project_Info'])
    for key, value in (project_info or {}).items():
        project_sheet.append([str(key), value])

    reviewer_sheet = workbook['Reviewer_Comments']
    reviewer_sheet.append(REVIEWER_COMMENT_HEADERS)
    for comment in validated_comments:
        reviewer_sheet.append(_reviewer_row(comment))
        highlight_cell = reviewer_sheet.cell(reviewer_sheet.max_row, 14)
        highlight_cell.fill = PatternFill(
            'solid', fgColor=_HIGHLIGHT_HEX[comment.highlight.value]
        )

    plan_sheet = workbook['Revision_Plan']
    plan_sheet.append(REVISION_PLAN_HEADERS)
    for sheet_name, headers in _OTHER_HEADERS.items():
        if sheet_name != 'Project_Info':
            workbook[sheet_name].append(headers)

    _build_dashboard(workbook)
    _write_enum_lists(workbook)
    _add_validations(workbook)

    header_map = {
        'Dashboard': ('Metric', 'Value'),
        'Project_Info': _OTHER_HEADERS['Project_Info'],
        'Reviewer_Comments': REVIEWER_COMMENT_HEADERS,
        'Revision_Plan': REVISION_PLAN_HEADERS,
        **{name: headers for name, headers in _OTHER_HEADERS.items() if name != 'Project_Info'},
    }
    for sheet_name, headers in header_map.items():
        sheet = workbook[sheet_name]
        if sheet_name != 'Dashboard':
            _format_header(sheet, headers)
        _format_body(sheet, headers)

    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    return destination


def _header_columns(worksheet: Any) -> dict[str, int]:
    return {
        str(cell.value): cell.column for cell in worksheet[1]
        if cell.value is not None
    }


def _assessment_values(assessment: GapAnalysisAssessment) -> dict[str, Any]:
    evidence = [
        f'{item.evidence_id}: {item.description}'
        for item in assessment.manuscript_evidence
    ]
    return {
        'Interpretation': assessment.interpretation,
        'Required Actions': _join(assessment.required_actions),
        'Target Sections': _join(assessment.target_sections),
        'Shared With': _join(assessment.shared_with_comments),
        'Coverage Status': (
            assessment.coverage_status.value if assessment.coverage_status else None
        ),
        'Manuscript Evidence': _join(evidence),
        'Missing Elements': _join(assessment.missing_elements),
        'Required References': _join(assessment.required_references),
        'Required Experiments': _join(assessment.required_experiments),
        'Required Statistics': _join(assessment.required_statistics),
        'Author Decision Required': assessment.author_decision_required,
        'Gap Analysis Confidence': assessment.confidence,
        'Manual Review Required': assessment.manual_review_required,
        'Evidence Status': (
            assessment.verification_status.value
            if assessment.verification_status else None
        ),
    }


def _action_row(action: RevisionAction) -> list[Any]:
    return [
        action.action_id,
        _join(action.comment_ids),
        action.change_type.value,
        action.target_section,
        action.target_object,
        action.old_text_summary,
        action.proposed_revision_summary or action.proposed_text,
        action.rationale,
        _join(action.evidence_ids),
        action.status.value,
        action.approval_state.value,
        action.highlight.value if action.highlight else None,
        action.applied_location,
        action.verified_by,
        action.verified_at.isoformat() if action.verified_at else None,
        _join(action.evidence_requirements),
        _join(action.reference_requirements),
        _join(action.experiment_requirements),
        _join(action.statistical_requirements),
        _join(action.unresolved_questions),
        action.approval_decision,
        action.author_note,
        action.modified_action_text,
        action.evidence_request,
        action.decision_timestamp.isoformat() if action.decision_timestamp else None,
        action.decision_maker,
    ]


def update_revision_workbook(
    path: str | Path,
    comments: Iterable[ReviewerComment],
    assessments: Iterable[GapAnalysisAssessment],
    actions: Iterable[RevisionAction],
    gate_status: ApprovalGateStatus | str,
) -> Path:
    '''Update the source-of-truth workbook without altering exact comment text.'''

    destination = Path(path)
    workbook = load_workbook(destination)
    reviewer = workbook['Reviewer_Comments']
    reviewer_headers = _header_columns(reviewer)
    if not set(REVIEWER_COMMENT_HEADERS).issubset(reviewer_headers):
        raise ValueError('Revision_Master.xlsx lacks Phase 4 Reviewer_Comments columns')
    source_comments = {comment.comment_id: comment for comment in comments}
    rows: dict[str, int] = {}
    for row in range(2, reviewer.max_row + 1):
        comment_id = reviewer.cell(row, reviewer_headers['Comment ID']).value
        if comment_id:
            rows[str(comment_id)] = row
    if set(rows) != set(source_comments):
        raise ValueError('workbook comment IDs do not match reviewer inventory')
    for comment_id, row in rows.items():
        stored = reviewer.cell(row, reviewer_headers['Original Comment']).value
        if stored != source_comments[comment_id].original_comment:
            raise ValueError(f'workbook changed exact comment text for {comment_id}')
    for assessment in assessments:
        if assessment.comment_id not in rows:
            raise ValueError(f'unknown workbook comment ID: {assessment.comment_id}')
        row = rows[assessment.comment_id]
        for header, value in _assessment_values(assessment).items():
            reviewer.cell(row, reviewer_headers[header], value)

    plan = workbook['Revision_Plan']
    plan_headers = _header_columns(plan)
    if not set(REVISION_PLAN_HEADERS).issubset(plan_headers):
        raise ValueError('Revision_Master.xlsx lacks Phase 4 Revision_Plan columns')
    if plan.max_row > 1:
        plan.delete_rows(2, plan.max_row - 1)
    validated_actions = [RevisionAction.model_validate(item) for item in actions]
    for row_number, action in enumerate(validated_actions, start=2):
        for column_number, value in enumerate(_action_row(action), start=1):
            plan.cell(row_number, column_number, value)
        plan._current_row = row_number
        plan.cell(plan.max_row, plan_headers['Highlight']).fill = PatternFill(
            'solid', fgColor=_HIGHLIGHT_HEX[action.highlight.value]
        )
    _format_body(reviewer, REVIEWER_COMMENT_HEADERS)
    _format_body(plan, REVISION_PLAN_HEADERS)
    reviewer.auto_filter.ref = (
        f'A1:{get_column_letter(len(REVIEWER_COMMENT_HEADERS))}{max(2, reviewer.max_row)}'
    )
    plan.auto_filter.ref = (
        f'A1:{get_column_letter(len(REVISION_PLAN_HEADERS))}{max(2, plan.max_row)}'
    )
    dashboard = workbook['Dashboard']
    status_value = gate_status.value if hasattr(gate_status, 'value') else str(gate_status)
    for row in range(1, dashboard.max_row + 1):
        if dashboard.cell(row, 1).value == 'Approval Gate status':
            dashboard.cell(row, 2, status_value)
            break
    workbook.save(destination)
    return destination


def _ensure_phase5_headers(worksheet: Any, headers: tuple[str, ...]) -> dict[str, int]:
    for column, header in enumerate(headers, start=1):
        worksheet.cell(1, column, header)
    _format_header(worksheet, headers)
    _format_body(worksheet, headers)
    return _header_columns(worksheet)


def update_revision_execution_workbook(
    path: str | Path,
    comments: Iterable[ReviewerComment],
    actions: Iterable[RevisionAction],
    drafts: Iterable[RevisionDraft],
    changes: Iterable[ChangeRecord],
    *,
    output_version: str | None,
    document_verification_status: str,
    blocked_change_count: int = 0,
) -> Path:
    '''Synchronize Phase 5 state without generating reviewer-response prose.'''

    destination = Path(path)
    workbook = load_workbook(destination)
    validated_comments = {
        item.comment_id: ReviewerComment.model_validate(item) for item in comments
    }
    validated_actions = [
        RevisionAction.model_validate(item) for item in actions
    ]
    validated_drafts = [
        RevisionDraft.model_validate(item) for item in drafts
    ]
    validated_changes = [
        ChangeRecord.model_validate(item) for item in changes
    ]

    reviewer = workbook['Reviewer_Comments']
    reviewer_headers = _ensure_phase5_headers(reviewer, REVIEWER_COMMENT_HEADERS)
    reviewer_rows: dict[str, int] = {}
    for row in range(2, reviewer.max_row + 1):
        value = reviewer.cell(row, reviewer_headers['Comment ID']).value
        if value:
            reviewer_rows[str(value)] = row
    if set(reviewer_rows) != set(validated_comments):
        raise ValueError('workbook comment IDs do not match reviewer inventory')
    for comment_id, row in reviewer_rows.items():
        if (
            reviewer.cell(row, reviewer_headers['Original Comment']).value
            != validated_comments[comment_id].original_comment
        ):
            raise ValueError(f'workbook changed exact comment text for {comment_id}')
        linked_actions = [
            action for action in validated_actions if comment_id in action.comment_ids
        ]
        linked_drafts = [
            draft for draft in validated_drafts if comment_id in draft.comment_ids
        ]
        linked_changes = [
            change for change in validated_changes if comment_id in change.comment_ids
        ]
        reviewer.cell(
            row, reviewer_headers['Approved Draft Count'],
            sum(draft.approval_state.value == 'APPROVED' for draft in linked_drafts),
        )
        reviewer.cell(
            row, reviewer_headers['Applied Change Count'], len(linked_changes)
        )
        applied_actions = {change.action_id for change in linked_changes}
        unresolved = sum(
            action.action_id not in applied_actions
            and action.approval_state.value != 'REJECTED'
            for action in linked_actions
        )
        reviewer.cell(
            row, reviewer_headers['Remaining Unresolved Actions'], unresolved
        )

    plan = workbook['Revision_Plan']
    plan_headers = _ensure_phase5_headers(plan, REVISION_PLAN_HEADERS)
    plan_rows = {
        str(plan.cell(row, plan_headers['Action ID']).value): row
        for row in range(2, plan.max_row + 1)
        if plan.cell(row, plan_headers['Action ID']).value
    }
    for action in validated_actions:
        if action.action_id not in plan_rows:
            raise ValueError(f'workbook is missing action {action.action_id}')
        row = plan_rows[action.action_id]
        linked = [
            draft for draft in validated_drafts if draft.action_id == action.action_id
        ]
        linked_changes = [
            change for change in validated_changes if change.action_id == action.action_id
        ]
        plan.cell(
            row, plan_headers['Draft Status'],
            _join(draft.draft_status for draft in linked),
        )
        plan.cell(
            row, plan_headers['Exact Text Approval State'],
            _join(draft.approval_state for draft in linked),
        )
        plan.cell(
            row, plan_headers['Application Status'],
            'APPLIED' if linked_changes else (
                _join(draft.application_status for draft in linked) or 'NOT_APPLIED'
            ),
        )
        plan.cell(
            row, plan_headers['Output Version'],
            output_version if linked_changes else None,
        )
        plan.cell(
            row, plan_headers['Verified Location'],
            _join(change.target_element_id for change in linked_changes),
        )

    change_sheet = workbook['Change_Log']
    change_headers = _ensure_phase5_headers(
        change_sheet, _OTHER_HEADERS['Change_Log']
    )
    if change_sheet.max_row > 1:
        change_sheet.delete_rows(2, change_sheet.max_row - 1)
    for record in validated_changes:
        values = {
            'Change ID': record.change_id,
            'Draft ID': record.draft_id,
            'Action ID': record.action_id,
            'Comment IDs': _join(record.comment_ids),
            'Operation': record.operation.value,
            'Target Section': record.target_section,
            'Target Element ID': record.target_element_id,
            'Old Text Hash': record.old_text_hash,
            'New Text Hash': record.new_text_hash,
            'Old Text Summary': record.old_text_summary,
            'New Text Summary': record.new_text_summary,
            'Highlight': record.highlight.value,
            'Applied At': record.application_timestamp.isoformat(),
            'Output Version': record.output_document_version,
            'Verification Status': record.verification_status,
            'Warnings': _join(record.warnings),
        }
        row = change_sheet.max_row + 1
        for header, value in values.items():
            change_sheet.cell(row, change_headers[header], value)

    response = workbook['Response_Map']
    response_headers = _header_columns(response)
    if response.max_row > 1:
        response.delete_rows(2, response.max_row - 1)
    for index, comment in enumerate(validated_comments.values(), start=1):
        linked = [
            change for change in validated_changes
            if comment.comment_id in change.comment_ids
        ]
        response.append([
            f'RESP-MAP-{index:04d}',
            comment.comment_id,
            True,
            None,
            _join(f'{change.change_id}:{change.operation.value}' for change in linked),
            _join(change.target_element_id for change in linked),
            'APPLIED' if linked else 'DRAFTED',
            comment.highlight.value,
            False,
            'Mapping prepared; final reviewer response has not been generated.',
        ])

    dashboard = workbook['Dashboard']
    metrics = {
        'Drafts prepared': len(validated_drafts),
        'Drafts awaiting text approval': sum(
            draft.approval_state.value == 'PENDING' for draft in validated_drafts
        ),
        'Text approvals completed': sum(
            draft.approval_state.value == 'APPROVED' for draft in validated_drafts
        ),
        'Changes applied': len(validated_changes),
        'Changes blocked': blocked_change_count,
        'Document verification status': document_verification_status,
    }
    dashboard_rows = {
        str(dashboard.cell(row, 1).value): row
        for row in range(1, dashboard.max_row + 1)
        if dashboard.cell(row, 1).value
    }
    for metric, value in metrics.items():
        row = dashboard_rows.get(metric)
        if row is None:
            row = dashboard.max_row + 1
            dashboard.cell(row, 1, metric)
        dashboard.cell(row, 2, value)

    _format_body(reviewer, REVIEWER_COMMENT_HEADERS)
    _format_body(plan, REVISION_PLAN_HEADERS)
    _format_body(change_sheet, _OTHER_HEADERS['Change_Log'])
    workbook.save(destination)
    return destination


def _append_missing_headers(worksheet: Any, headers: tuple[str, ...]) -> dict[str, int]:
    current = _header_columns(worksheet)
    for header in headers:
        if header not in current:
            worksheet.cell(1, worksheet.max_column + 1, header)
            current = _header_columns(worksheet)
    _format_header(worksheet, tuple(
        str(cell.value) for cell in worksheet[1] if cell.value is not None
    ))
    return current


def update_finalization_workbook(
    path: str | Path,
    package: ResponsePackage | dict,
    findings: Iterable[ConsistencyFinding | dict],
    readiness: str,
) -> Path:
    '''Synchronize Phase 7 fields without overwriting author notes.'''

    destination = Path(path)
    response_package = ResponsePackage.model_validate(package)
    consistency = [ConsistencyFinding.model_validate(item) for item in findings]
    workbook = load_workbook(destination)
    response = workbook['Response_Map']
    headers = _append_missing_headers(response, FINAL_RESPONSE_MAP_HEADERS)
    rows = {
        str(response.cell(row, headers['Comment ID']).value): row
        for row in range(2, response.max_row + 1)
        if response.cell(row, headers['Comment ID']).value
    }
    for entry in response_package.entries:
        row = rows.get(entry.comment_id)
        if row is None:
            row = response.max_row + 1
            response.cell(row, headers['Comment ID'], entry.comment_id)
            rows[entry.comment_id] = row
        values = {
            'Response Entry ID': entry.response_entry_id,
            'Response Status': entry.response_status.value,
            'Change IDs': _join(entry.related_change_ids),
            'Evidence IDs': _join(entry.related_evidence_ids),
            'Reference IDs': _join(entry.related_reference_ids),
            'Verified Location': _join(entry.verified_locations),
            'Highlight': entry.highlight.value,
            'Author Approval': entry.author_approved,
            'Verification Status': (
                'VERIFIED' if entry.response_status.value == 'VERIFIED'
                else entry.response_status.value
            ),
            'Unresolved Issues': _join([
                *entry.unresolved_limitations, *entry.verification_notes
            ]),
        }
        for header, value in values.items():
            response.cell(row, headers[header], value)
        legacy_values = {
            'Response ID': entry.response_entry_id,
            'Exact Comment Verified': True,
            'Response Draft': entry.author_response,
            'Changes Made': entry.changes_made,
            'Applied Location': _join(entry.verified_locations),
            'Status': entry.response_status.value,
            'Consistency Verified': not consistency,
        }
        for header, value in legacy_values.items():
            if header in headers:
                response.cell(row, headers[header], value)
        response.cell(row, headers['Highlight']).fill = PatternFill(
            'solid', fgColor=_HIGHLIGHT_HEX[entry.highlight.value]
        )
    dashboard = workbook['Dashboard']
    metric_rows = {
        str(dashboard.cell(row, 1).value): row
        for row in range(1, dashboard.max_row + 1)
        if dashboard.cell(row, 1).value
    }
    metrics = {
        'Total comments': len(response_package.entries),
        'Response drafts': sum(
            item.response_status.value in {'DRAFTED', 'AUTHOR_REVIEW'}
            for item in response_package.entries
        ),
        'Approved responses': sum(
            item.response_status.value in {'APPROVED', 'VERIFIED'}
            for item in response_package.entries
        ),
        'Verified responses': sum(
            item.response_status.value == 'VERIFIED'
            for item in response_package.entries
        ),
        'Blocked responses': sum(
            item.response_status.value == 'BLOCKED'
            for item in response_package.entries
        ),
        'Missing locations': sum(
            bool(item.related_change_ids) and not item.verified_locations
            for item in response_package.entries
        ),
        'Cross-document inconsistencies': len(consistency),
        'Final-release readiness': readiness,
    }
    for metric, value in metrics.items():
        row = metric_rows.get(metric)
        if row is None:
            row = dashboard.max_row + 1
            dashboard.cell(row, 1, metric)
            metric_rows[metric] = row
        dashboard.cell(row, 2, value)
    qa = workbook['QA_Findings']
    qa_headers = _header_columns(qa)
    existing = {
        str(qa.cell(row, qa_headers.get('Finding ID', 1)).value)
        for row in range(2, qa.max_row + 1)
    }
    for item in consistency:
        if item.finding_id in existing:
            continue
        qa.append([
            item.finding_id, 'RESPONSE_LETTER_CONSISTENCY',
            item.severity.value, item.description,
            '; '.join(item.documents), None, None, None,
            item.status.value, item.resolution, False,
        ])
    _format_body(response, tuple(
        str(cell.value) for cell in response[1] if cell.value is not None
    ))
    _format_body(qa, _OTHER_HEADERS['QA_Findings'])
    workbook.save(destination)
    return destination
