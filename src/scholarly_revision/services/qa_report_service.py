'''Aggregate Phase 6 findings, write reports, and synchronize the revision workbook.'''
from __future__ import annotations
import csv,json
from collections import Counter
from datetime import UTC,datetime
from pathlib import Path
from typing import Any,Iterable
from openpyxl import Workbook,load_workbook
from openpyxl.styles import Alignment,Font
from scholarly_revision.models.scientific_audit import (
    AuditIssue,AuditIssueStatus,AuditSeverity,FinalReleaseReadiness,ScientificQAReport,
)
from scholarly_revision.services.gap_analysis_service import read_json,write_json

ISSUE_HEADERS=('Issue ID','Category','Severity','Description','Document Element ID','Section',
 'Evidence','Related Comment IDs','Related Action IDs','Status','Resolution Required',
 'Resolution','Verified By','Verified At','Manual Review Required')
CSV_FIELDS=tuple(header.lower().replace(' ','_') for header in ISSUE_HEADERS)
EVIDENCE_WORDS=('evidence','source file','source range','registry','experiment','statistical')
SPECIAL_SHEETS={
 'REFERENCE':'Reference_Audit','CITATION':'Reference_Audit',
 'FIGURE_TABLE':'Figures_Tables','EQUATION_SYMBOL':'Notation_Equations',
 'NUMERICAL_CONSISTENCY':'Results_Integrity','RESULT_INTEGRITY':'Results_Integrity',
 'TERMINOLOGY':'Notation_Equations','HIGHLIGHT':'QA_Findings','FRONT_MATTER':'QA_Findings',
}

def readiness_for(issues:Iterable[AuditIssue])->FinalReleaseReadiness:
    items=list(issues); unresolved=[i for i in items if i.unresolved]
    if any(i.severity is AuditSeverity.BLOCKER for i in unresolved):
        return FinalReleaseReadiness.BLOCKED
    if any(i.severity is AuditSeverity.CRITICAL and
           any(word in i.description.casefold() for word in EVIDENCE_WORDS) for i in unresolved):
        return FinalReleaseReadiness.BLOCKED
    if any(i.severity in {AuditSeverity.CRITICAL,AuditSeverity.MAJOR} for i in unresolved):
        return FinalReleaseReadiness.NOT_READY
    if unresolved or items:
        return FinalReleaseReadiness.READY_WITH_WARNINGS
    return FinalReleaseReadiness.READY

def aggregate_report(results:Iterable[Any],source_hashes:dict[str,str]|None=None,
                     *,generated_at:datetime|None=None)->ScientificQAReport:
    result_list=list(results); issues=[issue for result in result_list for issue in result.issues]
    categories=Counter(i.category for i in issues); severities=Counter(i.severity.value for i in issues)
    statuses=Counter(i.status.value for i in issues); unresolved=[i for i in issues if i.unresolved]
    evidence=[i.issue_id for i in issues if any(word in i.description.casefold() for word in EVIDENCE_WORDS)]
    summaries={result.category:{
        'issue_count':len(result.issues),'checked_element_count':result.checked_element_count,
        'manual_review_required':result.manual_review_required,
    } for result in result_list}
    return ScientificQAReport(generated_at=generated_at or datetime.now(UTC),
        source_hashes=source_hashes or {},issues=issues,total_issues=len(issues),
        count_by_category=dict(sorted(categories.items())),
        count_by_severity=dict(sorted(severities.items())),
        count_by_status=dict(sorted(statuses.items())),
        blocker_count=sum(i.severity is AuditSeverity.BLOCKER and i.unresolved for i in issues),
        unresolved_critical_issues=sum(i.severity is AuditSeverity.CRITICAL and i.unresolved for i in issues),
        manual_review_count=sum(i.manual_review_required and i.unresolved for i in issues),
        final_release_readiness=readiness_for(issues),evidence_dependent_issue_ids=evidence,
        affected_sections=sorted({i.section for i in issues if i.section}),
        affected_objects=sorted({i.document_element_id for i in issues if i.document_element_id}),
        auditor_summaries=summaries)

def _issue_row(item:AuditIssue)->list[Any]:
    return [item.issue_id,item.category,item.severity.value,item.description,item.document_element_id,
        item.section,'; '.join(item.evidence),'; '.join(item.related_comment_ids),
        '; '.join(item.related_action_ids),item.status.value,item.resolution_required,item.resolution,
        item.verified_by,item.verified_at.isoformat() if item.verified_at else None,item.manual_review_required]

def _write_csv(path:Path,issues:list[AuditIssue])->None:
    with path.open('w',encoding='utf-8',newline='') as stream:
        writer=csv.writer(stream,lineterminator='\n');writer.writerow(ISSUE_HEADERS)
        writer.writerows(_issue_row(item) for item in issues)

def _write_xlsx(path:Path,report:ScientificQAReport)->None:
    workbook=Workbook(); summary=workbook.active;summary.title='Summary';summary.append(['Metric','Value'])
    metrics={'Total issues':report.total_issues,'Blocker count':report.blocker_count,
        'Unresolved critical issues':report.unresolved_critical_issues,
        'Manual review count':report.manual_review_count,
        'Final release readiness':report.final_release_readiness.value}
    metrics.update({f'Category: {k}':v for k,v in report.count_by_category.items()})
    metrics.update({f'Severity: {k}':v for k,v in report.count_by_severity.items()})
    for row in metrics.items():summary.append(row)
    issues=workbook.create_sheet('Issues');issues.append(ISSUE_HEADERS)
    for item in report.issues:issues.append(_issue_row(item))
    affected=workbook.create_sheet('Affected');affected.append(['Sections','Objects'])
    for index in range(max(len(report.affected_sections),len(report.affected_objects))):
        affected.append([report.affected_sections[index] if index<len(report.affected_sections) else None,
            report.affected_objects[index] if index<len(report.affected_objects) else None])
    for sheet in workbook.worksheets:
        for cell in sheet[1]:cell.font=Font(bold=True)
        sheet.freeze_panes='A2';sheet.auto_filter.ref=sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:cell.alignment=Alignment(vertical='top',wrap_text=True)
    workbook.save(path)

def decision_template(report:ScientificQAReport)->dict[str,Any]:
    return {'schema_version':1,'report_generated_at':report.generated_at.isoformat(),
        'instructions':'Record explicit decisions. Blank decisions are not resolutions.',
        'decisions':[{'issue_id':i.issue_id,'decision':None,'resolution':None,'justification':None,
            'evidence_request':None,'decision_maker':None,'decision_timestamp':None}
            for i in report.issues if i.unresolved]}

def write_qa_reports(project_root:str|Path,report:ScientificQAReport)->dict[str,Path]:
    root=Path(project_root).expanduser().resolve();audit=root/'audit';outputs=root/'outputs'
    audit.mkdir(parents=True,exist_ok=True);outputs.mkdir(parents=True,exist_ok=True)
    paths={'json':audit/'scientific_qa_report.json','csv':audit/'scientific_qa_report.csv',
        'xlsx':outputs/'Scientific_QA_Report.xlsx','decisions':audit/'qa_decision_template.json',
        'blockers':audit/'final_release_blockers.json'}
    write_json(paths['json'],report.model_dump(mode='json'));_write_csv(paths['csv'],report.issues)
    _write_xlsx(paths['xlsx'],report);write_json(paths['decisions'],decision_template(report))
    blockers=[i.model_dump(mode='json') for i in report.issues
        if i.unresolved and i.severity in {AuditSeverity.BLOCKER,AuditSeverity.CRITICAL}]
    write_json(paths['blockers'],{'readiness':report.final_release_readiness.value,
        'blocker_count':report.blocker_count,'issues':blockers})
    return paths

def _headers(sheet)->dict[str,int]:
    return {str(cell.value):cell.column for cell in sheet[1] if cell.value is not None}

def _ensure_headers(sheet,headers)->dict[str,int]:
    existing=_headers(sheet)
    for header in headers:
        if header not in existing:
            sheet.cell(1,sheet.max_column+1,header);existing[header]=sheet.max_column
    for cell in sheet[1]:cell.font=Font(bold=True)
    return existing

def _dashboard_set(sheet,metric,value)->None:
    for row in range(1,sheet.max_row+1):
        if sheet.cell(row,1).value==metric:sheet.cell(row,2,value);return
    sheet.append([metric,value])

def update_qa_workbook(path:str|Path,report:ScientificQAReport)->Path:
    destination=Path(path);workbook=load_workbook(destination)
    findings=workbook['QA_Findings'];columns=_ensure_headers(findings,ISSUE_HEADERS+('Author Notes',))
    existing={}
    for row in range(2,findings.max_row+1):
        key=findings.cell(row,columns['Issue ID']).value
        if key:existing[str(key)]={h:findings.cell(row,c).value for h,c in columns.items()}
    if findings.max_row>1:findings.delete_rows(2,findings.max_row-1)
    preserve=('Status','Resolution','Verified By','Verified At','Author Notes')
    for item in report.issues:
        values=dict(zip(ISSUE_HEADERS,_issue_row(item)));prior=existing.get(item.issue_id,{})
        for header in preserve:
            if prior.get(header) not in (None,''):values[header]=prior[header]
        row=findings.max_row+1
        for header,value in values.items():findings.cell(row,columns[header],value)
    dashboard=workbook['Dashboard'];unresolved=[i for i in report.issues if i.unresolved]
    mapping={'Open blockers':sum(i.severity is AuditSeverity.BLOCKER for i in unresolved),
        'Open critical issues':sum(i.severity is AuditSeverity.CRITICAL for i in unresolved),
        'Reference issues':sum(i.category in {'REFERENCE','CITATION'} for i in unresolved),
        'Numerical inconsistencies':sum(i.category=='NUMERICAL_CONSISTENCY' for i in unresolved),
        'Unverified results':sum(i.category=='RESULT_INTEGRITY' for i in unresolved),
        'Figure/table issues':sum(i.category=='FIGURE_TABLE' for i in unresolved),
        'Equation/symbol issues':sum(i.category=='EQUATION_SYMBOL' for i in unresolved),
        'Terminology issues':sum(i.category=='TERMINOLOGY' for i in unresolved),
        'Highlight issues':sum(i.category=='HIGHLIGHT' for i in unresolved),
        'Front-matter issues':sum(i.category=='FRONT_MATTER' for i in unresolved),
        'Final-release readiness':report.final_release_readiness.value}
    for key,value in mapping.items():_dashboard_set(dashboard,key,value)
    for category,sheet_name in SPECIAL_SHEETS.items():
        if sheet_name=='QA_Findings':continue
        sheet=workbook[sheet_name];cols=_ensure_headers(sheet,('QA Issue ID','QA Category','QA Severity',
            'QA Description','QA Status','QA Resolution','QA Evidence','QA Manual Review Required'))
        existing_ids={str(sheet.cell(r,cols['QA Issue ID']).value) for r in range(2,sheet.max_row+1)
                      if sheet.cell(r,cols['QA Issue ID']).value}
        for item in report.issues:
            if item.category!=category or item.issue_id in existing_ids:continue
            row=sheet.max_row+1;vals={'QA Issue ID':item.issue_id,'QA Category':item.category,
                'QA Severity':item.severity.value,'QA Description':item.description,
                'QA Status':item.status.value,'QA Resolution':item.resolution,
                'QA Evidence':'; '.join(item.evidence),'QA Manual Review Required':item.manual_review_required}
            for h,v in vals.items():sheet.cell(row,cols[h],v)
    response=workbook['Response_Map'];rcols=_ensure_headers(response,('QA Issue IDs','QA Open Count'))
    for row in range(2,response.max_row+1):
        comment=str(response.cell(row,2).value or '')
        linked=[i for i in report.issues if comment and comment in i.related_comment_ids]
        response.cell(row,rcols['QA Issue IDs'],'; '.join(i.issue_id for i in linked))
        response.cell(row,rcols['QA Open Count'],sum(i.unresolved for i in linked))
    change=workbook['Change_Log'];ccols=_ensure_headers(change,('QA Run At','QA Readiness'))
    row=change.max_row+1;change.cell(row,ccols['QA Run At'],report.generated_at.isoformat())
    change.cell(row,ccols['QA Readiness'],report.final_release_readiness.value)
    workbook.save(destination);return destination

def apply_qa_decisions(project_root:str|Path,decision_file:str|Path)->ScientificQAReport:
    root=Path(project_root).expanduser().resolve();report=ScientificQAReport.model_validate(
        read_json(root/'audit'/'scientific_qa_report.json'));payload=read_json(decision_file)
    raw=payload.get('decisions') if isinstance(payload,dict) else None
    if not isinstance(raw,list):raise ValueError('decision file must contain a decisions list')
    by_id={i.issue_id:i for i in report.issues};seen=set()
    for entry in raw:
        issue_id=str(entry.get('issue_id',''));decision=str(entry.get('decision') or '')
        if issue_id not in by_id:raise ValueError(f'unknown QA issue ID: {issue_id}')
        if issue_id in seen:raise ValueError(f'duplicate QA decision: {issue_id}')
        seen.add(issue_id);item=by_id[issue_id];maker=str(entry.get('decision_maker') or '').strip()
        timestamp=entry.get('decision_timestamp');when=datetime.fromisoformat(timestamp.replace('Z','+00:00')) if timestamp else datetime.now(UTC)
        if decision=='RESOLVE':
            resolution=str(entry.get('resolution') or '').strip()
            if not resolution:raise ValueError('RESOLVE requires a resolution description')
            update={'status':AuditIssueStatus.RESOLVED,'resolution':resolution,'verified_by':maker or None,'verified_at':when if maker else None}
        elif decision=='ACCEPT_RISK':
            justification=str(entry.get('justification') or '').strip()
            if not justification:raise ValueError('ACCEPT_RISK requires justification')
            if item.severity is AuditSeverity.BLOCKER and not maker:raise ValueError('BLOCKER ACCEPT_RISK requires an explicit decision maker')
            update={'status':AuditIssueStatus.ACCEPTED_RISK,'resolution':justification,'verified_by':maker or None,'verified_at':when if maker else None}
        elif decision=='NOT_APPLICABLE':
            resolution=str(entry.get('justification') or entry.get('resolution') or '').strip()
            if not resolution:raise ValueError('NOT_APPLICABLE requires justification')
            update={'status':AuditIssueStatus.NOT_APPLICABLE,'resolution':resolution,'verified_by':maker or None,'verified_at':when if maker else None}
        elif decision in {'DEFER','MANUAL_CORRECTION_REQUIRED'}:
            reason=str(entry.get('resolution') or entry.get('justification') or '').strip()
            if not reason:raise ValueError(f'{decision} requires a documented reason')
            update={'status':AuditIssueStatus.ACKNOWLEDGED,'resolution':f'{decision}: {reason}'}
        elif decision=='NEED_MORE_EVIDENCE':
            request=str(entry.get('evidence_request') or '').strip()
            if not request:raise ValueError('NEED_MORE_EVIDENCE requires an evidence request')
            update={'status':AuditIssueStatus.ACKNOWLEDGED,'resolution':f'Evidence requested: {request}'}
        else:raise ValueError(f'unsupported or blank QA decision for {issue_id}')
        by_id[issue_id]=item.model_copy(update=update)
    report=report.model_copy(update={'issues':[by_id[i.issue_id] for i in report.issues]})
    report=_reaggregate(report);write_qa_reports(root,report)
    workbook=root/'outputs'/'Revision_Master.xlsx'
    if workbook.is_file():update_qa_workbook(workbook,report)
    return report

def _reaggregate(report:ScientificQAReport)->ScientificQAReport:
    issues=report.issues;unresolved=[i for i in issues if i.unresolved]
    return report.model_copy(update={'total_issues':len(issues),
        'count_by_category':dict(sorted(Counter(i.category for i in issues).items())),
        'count_by_severity':dict(sorted(Counter(i.severity.value for i in issues).items())),
        'count_by_status':dict(sorted(Counter(i.status.value for i in issues).items())),
        'blocker_count':sum(i.severity is AuditSeverity.BLOCKER and i.unresolved for i in issues),
        'unresolved_critical_issues':sum(i.severity is AuditSeverity.CRITICAL and i.unresolved for i in issues),
        'manual_review_count':sum(i.manual_review_required and i.unresolved for i in issues),
        'final_release_readiness':readiness_for(issues)})

def verify_qa_resolutions(project_root:str|Path)->dict[str,Any]:
    root=Path(project_root).expanduser().resolve();report=ScientificQAReport.model_validate(
        read_json(root/'audit'/'scientific_qa_report.json'));errors=[]
    for item in report.issues:
        if item.status in {AuditIssueStatus.RESOLVED,AuditIssueStatus.ACCEPTED_RISK,AuditIssueStatus.NOT_APPLICABLE} and not (item.resolution or '').strip():
            errors.append(f'{item.issue_id} has no documented resolution')
        if item.status is AuditIssueStatus.RESOLVED and item.resolution_required and not item.verified_by:
            errors.append(f'{item.issue_id} resolution has no verifier')
    justified=readiness_for(report.issues)
    if justified!=report.final_release_readiness:errors.append('stored final-release readiness is not justified')
    result={'verified_at':datetime.now(UTC).isoformat(),'passed':not errors,'errors':errors,
        'final_release_readiness':justified.value,'blocker_count':sum(i.severity is AuditSeverity.BLOCKER and i.unresolved for i in report.issues)}
    write_json(root/'audit'/'qa_resolution_verification.json',result)
    if errors:raise ValueError('QA resolution verification failed: '+'; '.join(errors))
    return result
