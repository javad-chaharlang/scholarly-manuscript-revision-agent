import json
from pathlib import Path

import pytest
from openpyxl import load_workbook

from scholarly_revision.workflows.gap_analysis_workflow import import_and_plan, prepare_gap_analysis
from scholarly_revision.workflows.intake_workflow import IntakeRequest, run_intake_workflow

REVIEWERS = Path(__file__).parent / 'fixtures' / 'synthetic_reviewer_comments.docx'
MANUSCRIPT = Path(__file__).parent / 'fixtures' / 'synthetic_manuscript.docx'


def project(tmp_path: Path) -> Path:
    return run_intake_workflow(IntakeRequest(workspace_root=tmp_path / 'workspaces',
        project_name='Phase Four Synthetic', manuscript_id='SYNTHETIC',
        reviewer_file=REVIEWERS, manuscript_file=MANUSCRIPT)).workspace.root


def completed(root: Path) -> Path:
    payload = json.loads((root / 'working' / 'gap_analysis_template.json').read_text(encoding='utf-8'))
    for item in payload['assessments']:
        item.update({'coverage_status': 'NOT_ADDRESSED',
            'interpretation': 'Anonymous author-supplied interpretation.',
            'required_actions': ['Add a supported clarification.'],
            'target_sections': ['Introduction'], 'author_decision_required': True,
            'confidence': 0.7, 'manual_review_required': False})
    path = root / 'working' / 'completed.json'
    path.write_text(json.dumps(payload), encoding='utf-8')
    return path


def test_package_import_workbook_and_dashboard(tmp_path: Path) -> None:
    root = project(tmp_path)
    prepared = prepare_gap_analysis(root, MANUSCRIPT)
    package = json.loads(prepared.gap_analysis_input_path.read_text(encoding='utf-8'))
    assert prepared.comment_count == 7
    assert all(item['coverage_status'] is None for item in package['assessments'])
    result = import_and_plan(root, completed(root))
    assert result.action_count == 7
    workbook = load_workbook(result.workbook_path, data_only=False)
    assert workbook['Revision_Plan'].max_row == 8
    metrics = {workbook['Dashboard'].cell(r, 1).value: workbook['Dashboard'].cell(r, 2).value
        for r in range(1, workbook['Dashboard'].max_row + 1)}
    assert 'COUNTIF' in metrics['Fully addressed']
    assert metrics['Approval Gate status'] == 'READY_FOR_REVIEW'


def test_refuses_approved_plan_overwrite(tmp_path: Path) -> None:
    root = project(tmp_path); prepare_gap_analysis(root, MANUSCRIPT)
    analysis = completed(root); import_and_plan(root, analysis)
    path = root / 'working' / 'revision_plan.json'
    plan = json.loads(path.read_text(encoding='utf-8'))
    plan['actions'][0]['approval_state'] = 'APPROVED'
    path.write_text(json.dumps(plan), encoding='utf-8')
    with pytest.raises(ValueError, match='approved revision plan'):
        import_and_plan(root, analysis)
