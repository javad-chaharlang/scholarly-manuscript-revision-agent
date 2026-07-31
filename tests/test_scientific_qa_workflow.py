import json
from pathlib import Path
import pytest
from openpyxl import load_workbook
from phase6_helpers import CONFIG,REFERENCES,RESULTS,make_qa_project
from scholarly_revision.models.scientific_audit import AuditIssue
from scholarly_revision.services.qa_report_service import (
    aggregate_report,apply_qa_decisions,verify_qa_resolutions,
)
from scholarly_revision.workflows.scientific_qa_workflow import run_scientific_qa_workflow

class Result:
    category='TEST';checked_element_count=1;manual_review_required=False
    def __init__(self,issues):self.issues=issues

def test_report_aggregation_and_blocker_logic() -> None:
    issue=AuditIssue(issue_id='QA-TEST-0001',category='TEST',severity='BLOCKER',description='Missing evidence.')
    report=aggregate_report([Result([issue])])
    assert report.blocker_count==1
    assert report.final_release_readiness.value=='BLOCKED'

def test_workflow_reports_workbook_and_decisions(tmp_path:Path) -> None:
    root,highlighted,clean=make_qa_project(tmp_path)
    result=run_scientific_qa_workflow(project_root=root,highlighted_manuscript=highlighted,
        clean_manuscript=clean,results_registry=RESULTS,reference_registry=REFERENCES,config_path=CONFIG)
    assert result.report_paths['json'].is_file()
    workbook=load_workbook(result.workbook_path)
    assert workbook['QA_Findings'].max_row==result.report.total_issues+1
    metrics={workbook['Dashboard'].cell(r,1).value:workbook['Dashboard'].cell(r,2).value
        for r in range(1,workbook['Dashboard'].max_row+1)}
    assert metrics['Final-release readiness']=='BLOCKED'
    blocker=next(i for i in result.report.issues if i.severity.value=='BLOCKER')
    decisions={'decisions':[{'issue_id':blocker.issue_id,'decision':'RESOLVE',
        'resolution':'Anonymous synthetic evidence was documented.','decision_maker':'anonymous-verifier',
        'decision_timestamp':'2030-01-01T00:00:00Z'}]}
    path=root/'working'/'decisions.json';path.write_text(json.dumps(decisions),encoding='utf-8')
    updated=apply_qa_decisions(root,path)
    assert updated.blocker_count<result.report.blocker_count
    verify_qa_resolutions(root)

def test_decision_validation(tmp_path:Path) -> None:
    root,highlighted,clean=make_qa_project(tmp_path)
    run_scientific_qa_workflow(project_root=root,highlighted_manuscript=highlighted,clean_manuscript=clean)
    path=root/'working'/'bad.json';path.write_text(json.dumps({'decisions':[{
        'issue_id':'QA-HL-0001','decision':'ACCEPT_RISK'}]}),encoding='utf-8')
    with pytest.raises(ValueError,match='justification'):apply_qa_decisions(root,path)
