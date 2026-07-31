from pathlib import Path

from openpyxl import load_workbook

from scholarly_revision.tools.docx_reader import read_docx
from scholarly_revision.tools.reviewer_parser import parse_reviewer_comments
from scholarly_revision.tools.workbook_builder import (
    REVIEWER_COMMENT_HEADERS,
    REVISION_PLAN_HEADERS,
    REVISION_WORKBOOK_SHEETS,
    build_revision_workbook,
)


FIXTURE = Path(__file__).parent / 'fixtures' / 'synthetic_reviewer_comments.docx'


def make_workbook(tmp_path: Path):
    comments = parse_reviewer_comments(read_docx(FIXTURE)).comments
    path = build_revision_workbook(tmp_path / 'Revision_Master.xlsx', comments)
    return load_workbook(path, data_only=False)


def test_exact_sheet_names_and_header_rows(tmp_path: Path) -> None:
    workbook = make_workbook(tmp_path)
    assert tuple(workbook.sheetnames) == REVISION_WORKBOOK_SHEETS
    reviewer = workbook['Reviewer_Comments']
    plan = workbook['Revision_Plan']
    assert tuple(cell.value for cell in reviewer[1]) == REVIEWER_COMMENT_HEADERS
    assert tuple(cell.value for cell in plan[1]) == REVISION_PLAN_HEADERS
    assert reviewer.freeze_panes == 'A2'
    assert reviewer.auto_filter.ref.startswith('A1:R')


def test_workbook_uses_exact_policy_colors(tmp_path: Path) -> None:
    workbook = make_workbook(tmp_path)
    sheet = workbook['Reviewer_Comments']
    colors = {
        sheet.cell(row, 1).value: sheet.cell(row, 14).fill.fgColor.rgb
        for row in range(2, sheet.max_row + 1)
    }
    assert colors['R1-C01'] == 'FFFFFF00'
    assert colors['R2-C01'] == 'FF00FF00'
    assert colors['ED-C01'] == 'FFEE82EE'
    assert colors['GEN-C01'] == 'FFEE82EE'


def test_enum_data_validations_and_dashboard_formulas(tmp_path: Path) -> None:
    workbook = make_workbook(tmp_path)
    reviewer_validations = list(
        workbook['Reviewer_Comments'].data_validations.dataValidation
    )
    plan_validations = list(
        workbook['Revision_Plan'].data_validations.dataValidation
    )
    assert {item.formula1 for item in reviewer_validations} >= {
        '=CategoryValues', '=PriorityValues', '=StatusValues',
        '=EvidenceStatusValues', '=HighlightValues',
    }
    assert {item.formula1 for item in plan_validations} >= {
        '=ChangeTypeValues', '=StatusValues', '=ApprovalStateValues',
        '=HighlightValues',
    }
    dashboard = workbook['Dashboard']
    assert dashboard['B2'].value == '=COUNTA(Reviewer_Comments!A2:A1000)'
    assert 'COUNTIF' in dashboard['B6'].value
    assert dashboard['B7'].value.startswith('=IF(')
    assert dashboard.protection.sheet is True
