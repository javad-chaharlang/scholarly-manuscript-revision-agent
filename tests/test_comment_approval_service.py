from pathlib import Path

import pytest
from openpyxl import load_workbook

from phase5_helpers import (
    MANUSCRIPT,
    complete_and_import_drafts,
    decide_all_drafts,
    make_phase5_project,
    setup_approved_project,
)
from scholarly_revision.models.revision_draft import RevisionDraft
from scholarly_revision.models.project_state import ProjectState
from scholarly_revision.services.comment_approval_service import (
    eligible_draft_ids,
    validate_comment_approval_bundle,
)
from scholarly_revision.services.gap_analysis_service import read_json, write_json
from scholarly_revision.services.orchestrator_service import OrchestratorService
from scholarly_revision.services.project_state_service import ProjectStateService
from scholarly_revision.workflows.revision_execution_workflow import (
    apply_approved_revisions,
)


def test_complete_packet_covers_every_comment_and_only_exact_approved_drafts(
    tmp_path: Path,
) -> None:
    root = setup_approved_project(tmp_path)
    packet = read_json(root / 'working' / 'comment_approval_packet.json')
    bundle = validate_comment_approval_bundle(packet)
    comments = read_json(root / 'working' / 'reviewer_comments.json')
    assert {record.comment_id for record in bundle.records} == {
        item['comment_id'] for item in comments
    }
    drafts = [
        RevisionDraft.model_validate(item['draft'])
        for item in read_json(root / 'working' / 'revision_drafts.json')['drafts']
    ]
    assert eligible_draft_ids(bundle, drafts) == {
        'DRAFT-0001', 'DRAFT-0002', 'DRAFT-0003',
        'DRAFT-0004', 'DRAFT-0005', 'DRAFT-0006',
    }


def test_draft_edit_after_comment_approval_invalidates_authorization(
    tmp_path: Path,
) -> None:
    root = setup_approved_project(tmp_path)
    packet = validate_comment_approval_bundle(read_json(
        root / 'working' / 'comment_approval_packet.json'
    ))
    payload = read_json(root / 'working' / 'revision_drafts.json')
    payload['drafts'][0]['draft']['proposed_text'] = 'Changed after approval.'
    payload['drafts'][0]['draft']['approved_text'] = 'Changed after approval.'
    write_json(root / 'working' / 'revision_drafts.json', payload)
    drafts = [
        RevisionDraft.model_validate(item['draft']) for item in payload['drafts']
    ]
    assert 'DRAFT-0001' not in eligible_draft_ids(packet, drafts)


def test_manuscript_application_refuses_missing_comment_packet(tmp_path: Path) -> None:
    root = make_phase5_project(tmp_path)
    complete_and_import_drafts(root)
    decide_all_drafts(root)
    with pytest.raises(PermissionError, match='comment approval'):
        apply_approved_revisions(root, MANUSCRIPT)


def test_orchestrator_waits_for_all_comment_packages(tmp_path: Path) -> None:
    root = make_phase5_project(tmp_path)
    complete_and_import_drafts(root)
    decide_all_drafts(root)
    states = ProjectStateService(root)
    states.initialize(root.name, actor='anonymous-author')
    for target in (
        ProjectState.INTAKE_PENDING,
        ProjectState.GAP_ANALYSIS_PENDING,
        ProjectState.PLAN_APPROVAL,
        ProjectState.REVISION_DRAFTING,
        ProjectState.TEXT_APPROVAL,
    ):
        states.transition(target, action='synthetic-setup', actor='anonymous-author')
    orchestrator = OrchestratorService(tmp_path / 'private-workspaces')
    orchestrator.registry.register(
        root, ProjectState.TEXT_APPROVAL, project_id=root.name
    )
    assert not orchestrator.available_actions(root)['apply_revisions']
    orchestrator.prepare_comment_approval(root, actor='anonymous-author')
    approval = read_json(root / 'working' / 'comment_approval_working.json')
    state = states.load()
    for record in approval['records']:
        approved_ids = [
            item['draft_id'] for item in record.get('proposed_changes', [])
            if item.get('text_approval_state') == 'APPROVED'
            and not item.get('manual_handling_required')
        ]
        state = orchestrator.record_comment_approval(
            root,
            comment_id=record['comment_id'],
            proposed_response='Researcher-approved synthetic response.',
            decision='APPROVE_PACKAGE',
            decision_maker='anonymous-author',
            actor='anonymous-author',
            approved_draft_ids=approved_ids,
        )
    assert state.state is ProjectState.REVISION_APPLICATION
    assert orchestrator.available_actions(root)['apply_revisions']
    workbook = load_workbook(root / 'outputs' / 'Revision_Master.xlsx')
    response = workbook['Response_Map']
    headers = {cell.value: cell.column for cell in response[1]}
    assert 'Preapplication Decision' in headers
    assert 'Approved Draft IDs' in headers
    assert all(
        response.cell(row, headers['Preapplication Decision']).value
        == 'APPROVE_PACKAGE'
        for row in range(2, response.max_row + 1)
    )
