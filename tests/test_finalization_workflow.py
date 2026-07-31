from pathlib import Path

from openpyxl import load_workbook
import pytest

from phase7_helpers import make_ready_phase7_project
from scholarly_revision.workflows.finalization_workflow import (
    build_submission_package, generate_response_letter,
    run_final_consistency_check,
)
from scholarly_revision.services.project_workspace import sha256_file


def test_corrected_workflow_is_ready_and_updates_workbook(tmp_path: Path) -> None:
    root = make_ready_phase7_project(tmp_path)
    workbook_path = root / 'outputs' / 'Revision_Master.xlsx'
    inspected_hash = sha256_file(workbook_path)
    result = run_final_consistency_check(root)
    assert result.readiness == 'READY'
    assert result.consistency_json_path.is_file()
    assert result.consistency_csv_path.is_file()
    assert result.checklist_path.is_file()
    assert sha256_file(workbook_path) == inspected_hash
    workbook = load_workbook(workbook_path, data_only=True)
    response_headers = [cell.value for cell in workbook['Response_Map'][1]]
    assert 'Response Entry ID' in response_headers
    assert 'Unresolved Issues' in response_headers
    dashboard = {
        workbook['Dashboard'].cell(row, 1).value:
        workbook['Dashboard'].cell(row, 2).value
        for row in range(2, workbook['Dashboard'].max_row + 1)
    }
    assert dashboard['Verified responses'] == 1
    assert dashboard['Final-release readiness'] == 'READY'
    workbook.close()
    release = build_submission_package(root, 'release_v001')
    assert release.package_path.is_dir()


def test_false_scenario_blocks_workflow(tmp_path: Path) -> None:
    root = make_ready_phase7_project(tmp_path)
    (root / 'outputs' / 'Revised_Manuscript_Clean.docx').write_bytes(b'not a docx')
    result = run_final_consistency_check(root)
    assert result.readiness == 'BLOCKED'


def test_existing_response_letter_is_preserved(tmp_path: Path) -> None:
    root = make_ready_phase7_project(tmp_path)
    response = root / 'outputs' / 'Response_to_Reviewers.docx'
    package = root / 'working' / 'response_package.json'
    response_before = response.read_bytes()
    package_before = package.read_bytes()
    with pytest.raises(FileExistsError, match='already exists'):
        generate_response_letter(
            root, root / 'working' / 'completed_response.json'
        )
    assert response.read_bytes() == response_before
    assert package.read_bytes() == package_before
