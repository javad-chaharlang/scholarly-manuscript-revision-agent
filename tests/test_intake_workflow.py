import json
from hashlib import sha256
from pathlib import Path

import yaml
from openpyxl import load_workbook

from scholarly_revision.tools.workbook_builder import REVISION_WORKBOOK_SHEETS
from scholarly_revision.workflows.intake_workflow import (
    IntakeRequest,
    run_intake_workflow,
)


FIXTURE = Path(__file__).parent / 'fixtures' / 'synthetic_reviewer_comments.docx'


def test_intake_creates_private_artifacts_and_traceable_inventory(tmp_path: Path) -> None:
    result = run_intake_workflow(
        IntakeRequest(
            workspace_root=tmp_path / 'workspaces',
            project_name='Anonymous Intake Project',
            manuscript_id='SYNTHETIC-ID',
            reviewer_file=FIXTURE,
            reviewer_count=2,
        )
    )
    assert result.workspace.slug == 'anonymous-intake-project'
    for path in (
        result.manifest_path,
        result.reviewer_comments_path,
        result.intake_report_path,
        result.workbook_path,
    ):
        assert path.is_file()

    comments = json.loads(result.reviewer_comments_path.read_text(encoding='utf-8'))
    report = json.loads(result.intake_report_path.read_text(encoding='utf-8'))
    manifest_text = result.manifest_path.read_text(encoding='utf-8')
    manifest = yaml.safe_load(manifest_text)
    assert [item['comment_id'] for item in comments] == list(result.extracted_comment_ids)
    assert report['extracted_comment_count'] == 7
    assert report['manual_review_required_count'] == 1
    inventory = report['input_file_inventory'][0]
    assert inventory['sha256'] == sha256(FIXTURE.read_bytes()).hexdigest()
    assert (result.workspace.root / inventory['stored_path']).is_file()
    assert manifest['input_files']['reviewer_comments'] == [FIXTURE.name]
    assert 'Please clarify' not in manifest_text
    assert 'original_comment' not in manifest_text
    assert report['workspace_paths']['project_root'] == '.'
    assert 'manuscript_file (optional)' in report['missing_inputs']
    assert tuple(load_workbook(result.workbook_path).sheetnames) == REVISION_WORKBOOK_SHEETS


def test_optional_manuscript_is_copied_and_hashed_but_not_parsed(tmp_path: Path) -> None:
    manuscript = tmp_path / 'anonymous-manuscript.pdf'
    manuscript.write_bytes(b'%PDF-1.4\n% anonymous synthetic fixture\n')
    result = run_intake_workflow(
        IntakeRequest(
            workspace_root=tmp_path / 'workspaces',
            project_name='With Manuscript',
            manuscript_id='SYNTHETIC-ID',
            reviewer_file=FIXTURE,
            manuscript_file=manuscript,
        )
    )
    report = json.loads(result.intake_report_path.read_text(encoding='utf-8'))
    roles = {item['role'] for item in report['input_file_inventory']}
    assert roles == {'reviewer_comments', 'manuscript'}
    assert report['missing_inputs'] == []
