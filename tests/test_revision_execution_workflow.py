from hashlib import sha256
from pathlib import Path

from openpyxl import load_workbook

from phase5_helpers import MANUSCRIPT, setup_approved_project
from scholarly_revision.services.gap_analysis_service import read_json
from scholarly_revision.workflows.revision_execution_workflow import (
    apply_approved_revisions,
    verify_revision_outputs,
)


def test_end_to_end_applies_only_six_approved_texts(tmp_path: Path) -> None:
    root = setup_approved_project(tmp_path)
    source_before = sha256(MANUSCRIPT.read_bytes()).hexdigest()
    result = apply_approved_revisions(root, MANUSCRIPT)
    assert result.applied_change_count == 6
    assert result.blocked_change_count == 4
    assert result.output_version == 'v002'
    assert sha256(MANUSCRIPT.read_bytes()).hexdigest() == source_before
    report = verify_revision_outputs(root, MANUSCRIPT)
    assert report['passed']

    changes = read_json(root / 'audit' / 'change_log.json')['changes']
    assert len(changes) == 6
    assert {item['highlight'] for item in changes} == {
        'YELLOW', 'BRIGHT_GREEN', 'VIOLET'
    }
    workbook = load_workbook(root / 'outputs' / 'Revision_Master.xlsx')
    metrics = {
        workbook['Dashboard'].cell(row, 1).value:
        workbook['Dashboard'].cell(row, 2).value
        for row in range(1, workbook['Dashboard'].max_row + 1)
    }
    assert metrics['Changes applied'] == 6
    assert metrics['Changes blocked'] == 4
    assert workbook['Change_Log'].max_row == 7


def test_stale_approved_target_stops_without_output(tmp_path: Path) -> None:
    root = setup_approved_project(tmp_path)
    payload = read_json(root / 'working' / 'revision_drafts.json')
    payload['drafts'][0]['draft']['original_text_hash'] = 'f' * 64
    from scholarly_revision.services.gap_analysis_service import write_json
    write_json(root / 'working' / 'revision_drafts.json', payload)
    import pytest
    with pytest.raises(ValueError, match='changed after researcher approval'):
        apply_approved_revisions(root, MANUSCRIPT)
    assert not (root / 'outputs' / 'Revised_Manuscript_Highlighted.docx').exists()
